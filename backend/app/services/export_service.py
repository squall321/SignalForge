"""임원 보고용 데이터 export 서비스 (Track E).

기능
----
1) CSV  : voc_records 기간 추출 (csv 모듈, stdlib 만 사용 — 의존 없음)
2) Excel: voc_records + 요약 시트 (openpyxl)
3) PDF  : 보고서 (제목/KPI/타임라인/카테고리/키워드) — fpdf2

설계
----
- pandas 미사용 (개발 환경 최소화).  stdlib csv + openpyxl + fpdf2 만.
- DataFrame 추상화 없이 SQL → list[dict] → 파일 바이너리 직변환.
- 시리즈 코드 (예: GS25) 또는 시리즈 prefix (GS) 모두 허용 — `_resolve_product_ids`.
- period_days 기본 30, 최대 365.

테스트 가능성
-------------
이 모듈은 *AsyncSession* 을 주입받아 SQL 만 실행하므로 unit test 에서 in-memory
형태로 호출 가능.  endpoint 는 _internal 에서 thin wrapper.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession


# ── SQL ────────────────────────────────────────────────────────────────────
# 시리즈 코드(GS25) 또는 시리즈 prefix(GS) 모두 매칭.
_RESOLVE_PRODUCT_IDS_SQL = text(
    """
    SELECT id, code, name_en, name_ko, series_code
    FROM products
    WHERE is_active = TRUE
      AND (code = :series OR series_code = :series)
    ORDER BY released_at NULLS LAST, code
    """
)

# voc_records 추출 — 본문 200자 truncate 로 export 크기 제한.
_EXPORT_ROWS_SQL = text(
    """
    SELECT v.id,
           v.collected_at,
           v.published_at,
           p.code AS product_code,
           p.name_en AS product_name,
           pl.code AS platform_code,
           pl.name AS platform_name,
           v.country_code,
           v.language_detected,
           v.sentiment_score,
           v.sentiment_label,
           substr(v.content_original, 1, 200) AS content,
           v.source_url,
           v.likes_count,
           v.comments_count
    FROM voc_active v
    LEFT JOIN products p ON p.id = v.product_id
    LEFT JOIN platforms pl ON pl.id = v.platform_id
    WHERE v.product_id = ANY(:product_ids)
      AND v.collected_at >= now() - make_interval(days => :days)
    ORDER BY v.collected_at DESC
    LIMIT :limit
    """
)

# 요약 KPI — 시리즈 전체 집계.
_KPI_SQL = text(
    """
    SELECT count(*)::int AS total_voc,
           count(*) FILTER (WHERE sentiment_label = 'negative')::int AS neg_count,
           count(*) FILTER (WHERE sentiment_label = 'positive')::int AS pos_count,
           count(*) FILTER (WHERE sentiment_label = 'neutral')::int AS neu_count,
           round(avg(sentiment_score)::numeric, 4) AS avg_sent
    FROM voc_active
    WHERE product_id = ANY(:product_ids)
      AND collected_at >= now() - make_interval(days => :days)
    """
)

# 일별 타임라인 (max 90 일).
_TIMELINE_SQL = text(
    """
    SELECT date_trunc('day', collected_at)::date AS d,
           count(*)::int AS n,
           count(*) FILTER (WHERE sentiment_label = 'negative')::int AS neg
    FROM voc_active
    WHERE product_id = ANY(:product_ids)
      AND collected_at >= now() - make_interval(days => :days)
    GROUP BY d
    ORDER BY d
    """
)

# 카테고리 분포 (categories 는 TEXT[] — unnest)
_CATEGORIES_SQL = text(
    """
    SELECT unnest(categories) AS cat, count(*)::int AS n
    FROM voc_active
    WHERE product_id = ANY(:product_ids)
      AND collected_at >= now() - make_interval(days => :days)
      AND categories IS NOT NULL
    GROUP BY cat
    ORDER BY n DESC
    LIMIT :limit
    """
)

# 키워드 TOP-N (voc_keywords)
_KEYWORDS_SQL = text(
    """
    SELECT vk.keyword, count(*)::int AS n
    FROM voc_keywords vk
    JOIN voc_records v ON v.id = vk.voc_id
    WHERE v.product_id = ANY(:product_ids)
      AND v.collected_at >= now() - make_interval(days => :days)
    GROUP BY vk.keyword
    ORDER BY n DESC
    LIMIT :limit
    """
)


# ── 데이터 클래스 ───────────────────────────────────────────────────────────
@dataclass
class ExportContext:
    series: str
    period_days: int
    product_ids: List[int]
    product_codes: List[str]


# ── helpers ────────────────────────────────────────────────────────────────
async def _resolve_product_ids(
    db: AsyncSession, series: str
) -> Tuple[List[int], List[str]]:
    """series='GS25' (단일 product) 또는 'GS' (시리즈 prefix) → product_ids/codes."""
    rows = (await db.execute(_RESOLVE_PRODUCT_IDS_SQL, {"series": series})).all()
    if not rows:
        return [], []
    return [int(r.id) for r in rows], [r.code for r in rows]


async def build_context(
    db: AsyncSession, series: str, period_days: int
) -> ExportContext:
    period_days = max(1, min(int(period_days), 365))
    ids, codes = await _resolve_product_ids(db, series)
    if not ids:
        raise ValueError(f"unknown series/product: {series}")
    return ExportContext(
        series=series,
        period_days=period_days,
        product_ids=ids,
        product_codes=codes,
    )


async def _fetch_rows(
    db: AsyncSession, ctx: ExportContext, limit: int = 5000
) -> List[Dict[str, Any]]:
    rs = await db.execute(
        _EXPORT_ROWS_SQL,
        {"product_ids": ctx.product_ids, "days": ctx.period_days, "limit": int(limit)},
    )
    out: List[Dict[str, Any]] = []
    for r in rs:
        out.append({
            "id": int(r.id),
            "collected_at": r.collected_at.isoformat() if r.collected_at else "",
            "published_at": r.published_at.isoformat() if r.published_at else "",
            "product_code": r.product_code or "",
            "product_name": r.product_name or "",
            "platform_code": r.platform_code or "",
            "platform_name": r.platform_name or "",
            "country_code": r.country_code or "",
            "language": r.language_detected or "",
            "sentiment_score": (round(float(r.sentiment_score), 4)
                                if r.sentiment_score is not None else ""),
            "sentiment_label": r.sentiment_label or "",
            "content": (r.content or "").replace("\n", " ").replace("\r", " "),
            "source_url": r.source_url or "",
            "likes": int(r.likes_count or 0),
            "comments": int(r.comments_count or 0),
        })
    return out


async def _fetch_kpi(db: AsyncSession, ctx: ExportContext) -> Dict[str, Any]:
    r = (await db.execute(
        _KPI_SQL,
        {"product_ids": ctx.product_ids, "days": ctx.period_days},
    )).first()
    if r is None:
        return {"total_voc": 0, "neg_count": 0, "pos_count": 0,
                "neu_count": 0, "avg_sent": None, "neg_rate": 0.0}
    total = int(r.total_voc or 0)
    neg = int(r.neg_count or 0)
    return {
        "total_voc": total,
        "neg_count": neg,
        "pos_count": int(r.pos_count or 0),
        "neu_count": int(r.neu_count or 0),
        "avg_sent": (float(r.avg_sent) if r.avg_sent is not None else None),
        "neg_rate": (round(neg / total, 4) if total > 0 else 0.0),
    }


async def _fetch_timeline(
    db: AsyncSession, ctx: ExportContext
) -> List[Dict[str, Any]]:
    rs = await db.execute(
        _TIMELINE_SQL,
        {"product_ids": ctx.product_ids, "days": ctx.period_days},
    )
    return [
        {"date": r.d.isoformat(), "count": int(r.n), "negative": int(r.neg)}
        for r in rs
    ]


async def _fetch_categories(
    db: AsyncSession, ctx: ExportContext, limit: int = 10
) -> List[Dict[str, Any]]:
    rs = await db.execute(
        _CATEGORIES_SQL,
        {"product_ids": ctx.product_ids, "days": ctx.period_days, "limit": int(limit)},
    )
    return [{"category": r.cat, "count": int(r.n)} for r in rs]


async def _fetch_keywords(
    db: AsyncSession, ctx: ExportContext, limit: int = 20
) -> List[Dict[str, Any]]:
    rs = await db.execute(
        _KEYWORDS_SQL,
        {"product_ids": ctx.product_ids, "days": ctx.period_days, "limit": int(limit)},
    )
    return [{"keyword": r.keyword, "count": int(r.n)} for r in rs]


# ── CSV ────────────────────────────────────────────────────────────────────
def filename_for(ctx: ExportContext, ext: str) -> str:
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return f"voc_{ctx.series}_{today}.{ext}"


async def export_csv(db: AsyncSession, ctx: ExportContext) -> Tuple[bytes, str]:
    rows = await _fetch_rows(db, ctx)
    buf = io.StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    else:
        # 빈 결과여도 헤더는 출력 — 운영자가 schema 확인 가능
        writer = csv.DictWriter(
            buf,
            fieldnames=[
                "id", "collected_at", "published_at",
                "product_code", "product_name", "platform_code", "platform_name",
                "country_code", "language", "sentiment_score", "sentiment_label",
                "content", "source_url", "likes", "comments",
            ],
        )
        writer.writeheader()
    # UTF-8 BOM — Excel 호환
    data = ("﻿" + buf.getvalue()).encode("utf-8")
    return data, filename_for(ctx, "csv")


# ── Excel ──────────────────────────────────────────────────────────────────
async def export_excel(db: AsyncSession, ctx: ExportContext) -> Tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    rows = await _fetch_rows(db, ctx)
    kpi = await _fetch_kpi(db, ctx)
    timeline = await _fetch_timeline(db, ctx)
    categories = await _fetch_categories(db, ctx)
    keywords = await _fetch_keywords(db, ctx)

    wb = Workbook()
    # Sheet 1: 요약
    ws = wb.active
    ws.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E5E8F")
    ws["A1"] = "SignalForge VOC Export"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Series"
    ws["B3"] = ctx.series
    ws["A4"] = "Period (days)"
    ws["B4"] = ctx.period_days
    ws["A5"] = "Products"
    ws["B5"] = ", ".join(ctx.product_codes[:20])
    ws["A6"] = "Generated at (UTC)"
    ws["B6"] = datetime.now(timezone.utc).isoformat(timespec="seconds")

    ws["A8"] = "Total VOC"
    ws["B8"] = kpi["total_voc"]
    ws["A9"] = "Negative"
    ws["B9"] = kpi["neg_count"]
    ws["A10"] = "Positive"
    ws["B10"] = kpi["pos_count"]
    ws["A11"] = "Neutral"
    ws["B11"] = kpi["neu_count"]
    ws["A12"] = "Avg sentiment"
    ws["B12"] = kpi["avg_sent"] if kpi["avg_sent"] is not None else ""
    ws["A13"] = "Negative rate"
    ws["B13"] = kpi["neg_rate"]
    for r in range(3, 14):
        ws[f"A{r}"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    # Sheet 2: Timeline
    ws2 = wb.create_sheet("Timeline")
    ws2.append(["date", "count", "negative"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill
    for tl in timeline:
        ws2.append([tl["date"], tl["count"], tl["negative"]])

    # Sheet 3: Categories
    ws3 = wb.create_sheet("Categories")
    ws3.append(["category", "count"])
    for c in ws3[1]:
        c.font = header_font
        c.fill = header_fill
    for cat in categories:
        ws3.append([cat["category"], cat["count"]])

    # Sheet 4: Keywords
    ws4 = wb.create_sheet("Keywords")
    ws4.append(["keyword", "count"])
    for c in ws4[1]:
        c.font = header_font
        c.fill = header_fill
    for kw in keywords:
        ws4.append([kw["keyword"], kw["count"]])

    # Sheet 5: 원본 rows (5000 cap)
    ws5 = wb.create_sheet("VOC")
    if rows:
        headers = list(rows[0].keys())
        ws5.append(headers)
        for c in ws5[1]:
            c.font = header_font
            c.fill = header_fill
        for r in rows:
            ws5.append([r.get(h, "") for h in headers])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), filename_for(ctx, "xlsx")


# ── PDF ────────────────────────────────────────────────────────────────────
async def export_pdf(
    db: AsyncSession, ctx: ExportContext, sections: List[str]
) -> Tuple[bytes, str]:
    """PDF 보고서. sections: kpi/timeline/categories/keywords 중 선택.

    fpdf2 의 core font (Helvetica) 만 사용 — 한글 미지원이라 한국어는 *로 대체.
    임원 보고용 영문 fallback 으로 충분 (시리즈 코드, 숫자, 영문 라벨).
    """
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    kpi = await _fetch_kpi(db, ctx) if "kpi" in sections else None
    timeline = await _fetch_timeline(db, ctx) if "timeline" in sections else None
    categories = (
        await _fetch_categories(db, ctx) if "categories" in sections else None
    )
    keywords = await _fetch_keywords(db, ctx) if "keywords" in sections else None

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    def _safe(s: str) -> str:
        # fpdf core font 는 latin-1 만 — 한글/이모지 제거
        try:
            s.encode("latin-1")
            return s
        except UnicodeEncodeError:
            return s.encode("latin-1", "replace").decode("latin-1")

    def _ln_cell(w: float, h: float, txt: str, border: int = 0) -> None:
        """줄바꿈 cell (deprecated ``ln=1`` 대신 새 API 사용)."""
        pdf.cell(w, h, txt, border=border,
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.set_font("Helvetica", "B", 18)
    _ln_cell(0, 10, _safe("SignalForge VOC Report"))
    pdf.set_font("Helvetica", "", 10)
    _ln_cell(
        0, 6,
        _safe(
            f"Series: {ctx.series}   Period: {ctx.period_days}d   "
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        ),
    )
    _ln_cell(0, 6, _safe(f"Products: {', '.join(ctx.product_codes[:15])}"))
    pdf.ln(4)

    if kpi:
        pdf.set_font("Helvetica", "B", 13)
        _ln_cell(0, 8, "KPI Summary")
        pdf.set_font("Helvetica", "", 10)
        for label, key in [
            ("Total VOC", "total_voc"),
            ("Negative", "neg_count"),
            ("Positive", "pos_count"),
            ("Neutral", "neu_count"),
            ("Avg sentiment", "avg_sent"),
            ("Negative rate", "neg_rate"),
        ]:
            val = kpi.get(key)
            pdf.cell(60, 6, f"  {label}", border=0)
            _ln_cell(0, 6, str(val if val is not None else "-"))
        pdf.ln(3)

    if timeline:
        pdf.set_font("Helvetica", "B", 13)
        _ln_cell(0, 8, "Timeline (last entries)")
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(40, 6, "Date", border=1)
        pdf.cell(30, 6, "Count", border=1)
        _ln_cell(30, 6, "Negative", border=1)
        pdf.set_font("Helvetica", "", 9)
        for tl in timeline[-30:]:
            pdf.cell(40, 5, tl["date"], border=1)
            pdf.cell(30, 5, str(tl["count"]), border=1)
            _ln_cell(30, 5, str(tl["negative"]), border=1)
        pdf.ln(3)

    if categories:
        pdf.set_font("Helvetica", "B", 13)
        _ln_cell(0, 8, "Top categories")
        pdf.set_font("Helvetica", "", 10)
        for c in categories[:10]:
            _ln_cell(0, 5, _safe(f"  - {c['category']}: {c['count']}"))
        pdf.ln(3)

    if keywords:
        pdf.set_font("Helvetica", "B", 13)
        _ln_cell(0, 8, "Top keywords")
        pdf.set_font("Helvetica", "", 10)
        for k in keywords[:20]:
            _ln_cell(0, 5, _safe(f"  - {k['keyword']}: {k['count']}"))

    # fpdf2 returns bytearray; convert to bytes
    out = pdf.output()
    if isinstance(out, (bytearray, bytes)):
        data = bytes(out)
    else:
        # fallback (older API returns str/None)
        data = bytes(out, "latin-1") if isinstance(out, str) else b""
    return data, filename_for(ctx, "pdf")


# ── 공유 토큰 (in-memory, TTL) ─────────────────────────────────────────────
# 운영용 localhost only 환경.  단순 dict + TTL 검증.
# 재시작 시 토큰 무효화 — 임시 공유에 적합.
import secrets
import threading
import time

_TOKEN_LOCK = threading.Lock()
_TOKENS: Dict[str, Dict[str, Any]] = {}


def create_share_token(resource: str, expires_in: int) -> Dict[str, Any]:
    """공유 토큰 발급.  expires_in 단위 = 초 (max 30일)."""
    if not resource or not resource.startswith("/"):
        raise ValueError("resource must start with '/'")
    expires_in = max(60, min(int(expires_in), 30 * 86400))
    token = secrets.token_urlsafe(16)
    now = int(time.time())
    rec = {
        "token": token,
        "resource": resource,
        "created_at": now,
        "expires_at": now + expires_in,
    }
    with _TOKEN_LOCK:
        # cleanup expired
        for k in list(_TOKENS.keys()):
            if _TOKENS[k]["expires_at"] < now:
                _TOKENS.pop(k, None)
        _TOKENS[token] = rec
    return {
        "token": token,
        "url": f"/shared/{token}",
        "resource": resource,
        "expires_at": datetime.fromtimestamp(
            rec["expires_at"], tz=timezone.utc
        ).isoformat(timespec="seconds"),
    }


def resolve_share_token(token: str) -> Optional[Dict[str, Any]]:
    """토큰 검증.  만료/미존재 시 None."""
    now = int(time.time())
    with _TOKEN_LOCK:
        rec = _TOKENS.get(token)
        if not rec:
            return None
        if rec["expires_at"] < now:
            _TOKENS.pop(token, None)
            return None
        return {
            "token": token,
            "resource": rec["resource"],
            "expires_at": datetime.fromtimestamp(
                rec["expires_at"], tz=timezone.utc
            ).isoformat(timespec="seconds"),
        }


def share_tokens_stats() -> Dict[str, Any]:
    """디버그용 — 현재 활성 토큰 개수."""
    now = int(time.time())
    with _TOKEN_LOCK:
        active = sum(1 for r in _TOKENS.values() if r["expires_at"] >= now)
    return {"active": active}
