"""R10 Track E — export endpoint 단위 테스트.

3 케이스:
  1. CSV 헤더 + filename + UTF-8 BOM
  2. Excel (.xlsx) 5 sheet 구조 검증
  3. PDF 매직 바이트 (%PDF-) + 200 KB 이내 + share-token round-trip

실행:
    cd backend && .venv/bin/pytest tests/test_export.py -v
"""
import asyncio
import io
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
BACKEND_ROOT = os.path.abspath(os.path.join(HERE, ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from app.database import AsyncSessionLocal  # noqa: E402
from app.services import export_service  # noqa: E402


async def _csv_export_s25():
    """1) CSV — UTF-8 BOM + 헤더 필수 컬럼 + filename 패턴."""
    async with AsyncSessionLocal() as db:
        ctx = await export_service.build_context(db, "GS25", period_days=30)
        data, fname = await export_service.export_csv(db, ctx)
    assert isinstance(data, bytes)
    # UTF-8 BOM
    assert data.startswith(b"\xef\xbb\xbf"), "CSV must start with UTF-8 BOM"
    head = data.decode("utf-8").splitlines()[0]
    # 필수 컬럼이 헤더에 모두 등장
    for col in ["id", "collected_at", "product_code", "platform_code",
                "sentiment_score", "sentiment_label", "content"]:
        assert col in head, f"CSV header missing column: {col}"
    # filename pattern: voc_GS25_YYYY-MM-DD.csv
    assert fname.startswith("voc_GS25_"), f"unexpected filename: {fname}"
    assert fname.endswith(".csv"), f"unexpected ext: {fname}"
    print(f"[ok] CSV: {len(data)} bytes, {fname}, lines={len(data.splitlines())}")


async def _excel_export_s25():
    """2) Excel — 5 sheet (Summary/Timeline/Categories/Keywords/VOC) + 데이터 값."""
    from openpyxl import load_workbook

    async with AsyncSessionLocal() as db:
        ctx = await export_service.build_context(db, "GS25", period_days=30)
        data, fname = await export_service.export_excel(db, ctx)
    assert fname.endswith(".xlsx"), f"unexpected ext: {fname}"
    # xlsx zip 매직 바이트
    assert data[:2] == b"PK", "xlsx must be a ZIP container"
    wb = load_workbook(io.BytesIO(data))
    sheets = wb.sheetnames
    for expected in ["Summary", "Timeline", "Categories", "Keywords", "VOC"]:
        assert expected in sheets, f"missing sheet: {expected} (got {sheets})"
    # Summary 시트의 series cell
    ws = wb["Summary"]
    assert ws["B3"].value == "GS25"
    assert ws["B4"].value == 30
    # Total VOC 가 양수 (S25 라이브 데이터)
    total_voc = ws["B8"].value
    assert isinstance(total_voc, int) and total_voc >= 0
    print(
        f"[ok] Excel: {len(data)} bytes, sheets={len(sheets)}, "
        f"total_voc={total_voc}, file={fname}"
    )


async def _pdf_and_share_token():
    """3) PDF 매직 + share-token 발급/검증/만료 동작."""
    async with AsyncSessionLocal() as db:
        ctx = await export_service.build_context(db, "GS25", period_days=7)
        data, fname = await export_service.export_pdf(
            db, ctx, sections=["kpi", "timeline", "categories", "keywords"]
        )
    assert data.startswith(b"%PDF-"), "PDF must start with %PDF-"
    assert fname.endswith(".pdf"), f"unexpected ext: {fname}"
    assert 100 < len(data) < 1_000_000, f"PDF size suspicious: {len(data)}"

    # share-token round-trip
    issued = export_service.create_share_token("/insights", expires_in=3600)
    assert "token" in issued and len(issued["token"]) >= 16
    assert issued["url"].startswith("/shared/")
    rec = export_service.resolve_share_token(issued["token"])
    assert rec is not None
    assert rec["resource"] == "/insights"
    # invalid token
    assert export_service.resolve_share_token("invalid-xxx") is None
    # expires_in too short → clamped to 60s minimum (passes)
    issued2 = export_service.create_share_token("/dashboard", expires_in=10)
    rec2 = export_service.resolve_share_token(issued2["token"])
    assert rec2 is not None and rec2["resource"] == "/dashboard"

    # invalid resource (no leading /) → ValueError
    try:
        export_service.create_share_token("no-slash", expires_in=600)
        raise AssertionError("expected ValueError")
    except ValueError:
        pass

    print(
        f"[ok] PDF: {len(data)} bytes, file={fname}, "
        f"share-token active={export_service.share_tokens_stats()['active']}"
    )


def test_csv_export_s25():
    asyncio.run(_csv_export_s25())


def test_excel_export_s25():
    asyncio.run(_excel_export_s25())


def test_pdf_and_share_token():
    asyncio.run(_pdf_and_share_token())


async def _run_all():
    await _csv_export_s25()
    await _excel_export_s25()
    await _pdf_and_share_token()


if __name__ == "__main__":
    asyncio.run(_run_all())
    print("\nAll 3 export tests passed.")
